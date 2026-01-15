#!/usr/bin/env python3
"""
VNSH eSIM Inventory Import Tool
Imports eSIMs from Excel (.xlsx) or CSV files.

Usage:
    python3 import_esims.py <file.xlsx|file.csv> [--clear]

Options:
    --clear     Clear existing inventory before importing

Expected columns (case-insensitive, flexible naming):
    - lpa / lpa_code / activation_code / code
    - country / destination / region
    - days / duration / validity
    - price / price_usd / cost
"""

import sys
import os
from pathlib import Path

# Add parent to path for database import
sys.path.insert(0, str(Path(__file__).parent))

from database import init_db, get_db, get_inventory_stats, delete_all_inventory

def normalize_column_name(col):
    """Normalize column name to expected format."""
    col = col.lower().strip()

    # LPA mappings
    if col in ['lpa', 'lpa_code', 'activation_code', 'code', 'esim_code', 'qr_code', 'activation']:
        return 'lpa'

    # Country mappings
    if col in ['country', 'destination', 'region', 'location', 'country_code']:
        return 'country'

    # Days mappings
    if col in ['days', 'duration', 'validity', 'duration_days', 'valid_days', 'period']:
        return 'days'

    # GB/Data mappings
    if col in ['gb', 'data_gb', 'data', 'gigabytes', 'data_amount']:
        return 'gb'

    # Price mappings
    if col in ['price', 'price_usd', 'cost', 'usd', 'amount', 'price_usd']:
        return 'price'

    return col

def import_excel(filepath):
    """Import eSIMs from Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0, ["Excel file is empty"]

    # Get headers from first row
    raw_headers = rows[0]
    headers = [normalize_column_name(str(h)) if h else '' for h in raw_headers]

    # Check required columns
    required = ['lpa', 'country', 'days', 'price']
    missing = [r for r in required if r not in headers]
    if missing:
        return 0, 0, [f"Missing required columns: {', '.join(missing)}. Found: {', '.join([str(h) for h in raw_headers if h])}"]

    conn = get_db()
    cursor = conn.cursor()

    imported = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            # Build dict from row
            data = dict(zip(headers, row))

            lpa = str(data.get('lpa', '') or '').strip()
            country = str(data.get('country', '') or '').strip()
            days = data.get('days')
            gb = data.get('gb', 1)  # Default to 1GB
            price = data.get('price')

            # Validate
            if not lpa:
                errors.append(f"Row {row_num}: Missing LPA")
                skipped += 1
                continue
            if not country:
                errors.append(f"Row {row_num}: Missing country")
                skipped += 1
                continue
            if days is None:
                errors.append(f"Row {row_num}: Missing days")
                skipped += 1
                continue
            if price is None:
                errors.append(f"Row {row_num}: Missing price")
                skipped += 1
                continue

            # Convert types
            try:
                days = int(float(days))
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Invalid days value '{days}'")
                skipped += 1
                continue

            try:
                gb = int(float(gb)) if gb else 1
                # Sanity check - if GB is unreasonably large, default to 1
                if gb > 1000:
                    gb = 1
            except (ValueError, TypeError):
                gb = 1

            try:
                price = float(price)
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Invalid price value '{price}'")
                skipped += 1
                continue

            # Insert
            cursor.execute("""
                INSERT INTO esim_inventory (lpa, country, days, data_gb, price_usd, status)
                VALUES (?, ?, ?, ?, ?, 'available')
            """, (lpa, country, days, gb, price))
            imported += 1

        except Exception as e:
            if 'UNIQUE constraint' in str(e):
                errors.append(f"Row {row_num}: Duplicate LPA '{lpa}'")
            else:
                errors.append(f"Row {row_num}: {str(e)}")
            skipped += 1

    conn.commit()
    conn.close()
    wb.close()

    return imported, skipped, errors

def import_csv(filepath):
    """Import eSIMs from CSV file."""
    import csv

    conn = get_db()
    cursor = conn.cursor()

    imported = 0
    skipped = 0
    errors = []

    with open(filepath, 'r', newline='', encoding='utf-8') as f:
        # Try to detect delimiter
        sample = f.read(4096)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)

        # Normalize headers
        if reader.fieldnames:
            header_map = {h: normalize_column_name(h) for h in reader.fieldnames}
        else:
            return 0, 0, ["CSV file has no headers"]

        # Check required columns
        required = ['lpa', 'country', 'days', 'price']
        found_cols = set(header_map.values())
        missing = [r for r in required if r not in found_cols]
        if missing:
            return 0, 0, [f"Missing required columns: {', '.join(missing)}. Found: {', '.join(reader.fieldnames)}"]

        for row_num, row in enumerate(reader, start=2):
            try:
                # Normalize row keys
                data = {header_map[k]: v for k, v in row.items() if k in header_map}

                lpa = str(data.get('lpa', '') or '').strip()
                country = str(data.get('country', '') or '').strip()
                days = data.get('days', '').strip() if data.get('days') else ''
                gb = data.get('gb', '').strip() if data.get('gb') else '1'
                price = data.get('price', '').strip() if data.get('price') else ''

                # Validate
                if not lpa:
                    errors.append(f"Row {row_num}: Missing LPA")
                    skipped += 1
                    continue
                if not country:
                    errors.append(f"Row {row_num}: Missing country")
                    skipped += 1
                    continue
                if not days:
                    errors.append(f"Row {row_num}: Missing days")
                    skipped += 1
                    continue
                if not price:
                    errors.append(f"Row {row_num}: Missing price")
                    skipped += 1
                    continue

                # Convert types
                try:
                    days = int(float(days))
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid days value '{days}'")
                    skipped += 1
                    continue

                try:
                    gb = int(float(gb)) if gb else 1
                    # Sanity check - if GB is unreasonably large, default to 1
                    if gb > 1000:
                        gb = 1
                except ValueError:
                    gb = 1

                try:
                    price = float(price)
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid price value '{price}'")
                    skipped += 1
                    continue

                # Insert
                cursor.execute("""
                    INSERT INTO esim_inventory (lpa, country, days, data_gb, price_usd, status)
                    VALUES (?, ?, ?, ?, ?, 'available')
                """, (lpa, country, days, gb, price))
                imported += 1

            except Exception as e:
                if 'UNIQUE constraint' in str(e):
                    errors.append(f"Row {row_num}: Duplicate LPA '{lpa}'")
                else:
                    errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1

    conn.commit()
    conn.close()

    return imported, skipped, errors


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = sys.argv[1]
    clear_first = '--clear' in sys.argv

    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    # Initialize database
    init_db()

    # Show current stats
    print("Current inventory:")
    stats = get_inventory_stats()
    for status, count in stats.items():
        print(f"  {status}: {count}")
    print()

    # Clear if requested
    if clear_first:
        deleted = delete_all_inventory()
        print(f"Cleared {deleted} existing eSIMs\n")

    # Determine file type and import
    ext = Path(filepath).suffix.lower()

    if ext in ['.xlsx', '.xls']:
        print(f"Importing from Excel: {filepath}")
        imported, skipped, errors = import_excel(filepath)
    elif ext == '.csv':
        print(f"Importing from CSV: {filepath}")
        imported, skipped, errors = import_csv(filepath)
    else:
        print(f"Error: Unsupported file type: {ext}")
        print("Supported formats: .xlsx, .xls, .csv")
        sys.exit(1)

    # Report results
    print(f"\nImport complete:")
    print(f"  Imported: {imported}")
    print(f"  Skipped:  {skipped}")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:20]:  # Show first 20 errors
            print(f"  - {err}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more errors")

    # Show new stats
    print("\nNew inventory:")
    stats = get_inventory_stats()
    for status, count in stats.items():
        print(f"  {status}: {count}")


if __name__ == "__main__":
    main()
